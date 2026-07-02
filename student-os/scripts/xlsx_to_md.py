#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from itertools import islice
from pathlib import Path


def yaml_string(value: str) -> str:
    return json.dumps(value)


def load_workbook_module():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency 'openpyxl'. Install the packages from requirements.txt before running this script."
        ) from exc
    return load_workbook


def stringify(value: object) -> str:
    return "" if value is None else str(value)


def markdown_cell(value: object) -> str:
    return stringify(value).replace("|", "\\|").replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert an XLSX workbook into a markdown table summary.")
    parser.add_argument("xlsx", help="Path to the XLSX file")
    parser.add_argument("--output", required=True, help="Output markdown path")
    parser.add_argument("--max-rows", type=int, default=12, help="Maximum rows per sheet to render")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    load_workbook = load_workbook_module()
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    formula_wb = load_workbook(filename=str(xlsx_path), data_only=False)
    lines = [
        "---",
        "type: imported-table-summary",
        "course:",
        "status: active",
        "created:",
        "updated:",
        "tags: [import, table]",
        f"source_file: {yaml_string(str(xlsx_path))}",
        "import_method: xlsx-to-md",
        "repair_status:",
        "derived_from_import:",
        "---",
        "",
        f"# Imported Table Summary - {xlsx_path.stem}",
        "",
        "## Source",
        "",
        f"- Source file: {xlsx_path}",
        "- Import method: xlsx-to-md",
        "",
        "## Tables",
        "",
    ]

    for ws, formula_ws in zip(wb.worksheets, formula_wb.worksheets):
        lines.extend([f"## Sheet: {ws.title}", ""])
        rows = list(islice(ws.iter_rows(values_only=False), args.max_rows))
        formula_rows = list(islice(formula_ws.iter_rows(values_only=False), args.max_rows))
        if not rows:
            lines.extend(["- Empty sheet", ""])
            continue
        header = [markdown_cell(cell.value) for cell in rows[0]]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row, formula_row in zip(rows[1:], formula_rows[1:]):
            padded = list(row[: len(header)]) + ([None] * max(0, len(header) - len(row)))
            formula_padded = list(formula_row[: len(header)]) + ([None] * max(0, len(header) - len(formula_row)))
            values = []
            for cell, formula_cell in zip(padded, formula_padded):
                cell_value = None if cell is None else cell.value
                formula_value = None if formula_cell is None else formula_cell.value
                if cell_value is None and isinstance(formula_value, str) and formula_value.startswith("="):
                    values.append(markdown_cell(formula_value))
                else:
                    values.append(markdown_cell(cell_value))
            lines.append("| " + " | ".join(values) + " |")
        if ws.max_row > args.max_rows:
            lines.extend(["", f"- Truncated after {args.max_rows} rows.", ""])
        else:
            lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(json.dumps({"output": str(output_path)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
